from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from chinese_calendar import is_workday

from app.config import EXPORT_DIR, ensure_directories, resolve_target
from app.db import Database
from app.rules import DOC_NUMBER_RE, WorkdayCalendar, normalize_document_number


HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
ALERT_FILL = PatternFill("solid", fgColor="FCE4D6")


def _excel_date(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(value).date()
    except (TypeError, ValueError):
        return value


def _workday_count(detail: str):
    match = re.search(r"相隔\s*(\d+)\s*个工作日", detail or "")
    return int(match.group(1)) if match else None


def _document_number_parts(value: str) -> tuple[str, int | None, int | None, str]:
    normalized = normalize_document_number(value or "")
    match = DOC_NUMBER_RE.search(normalized)
    if not match:
        return "", None, None, normalized
    return match.group(1), int(match.group(2)), int(match.group(3)), normalized


def _timeliness(authored, published) -> tuple[int | None, str]:
    if not authored or not published:
        return None, "待复核"
    if published < authored:
        return 0, "日期倒置"
    try:
        count = WorkdayCalendar(provider=is_workday).count_between(authored, published)
    except NotImplementedError:
        count = WorkdayCalendar().count_between(authored, published)
    return count, "超期" if count > 20 else "合规"


def _format_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = HEAD_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in range(1, sheet.max_column + 1):
        values = [str(sheet.cell(row, column).value or "") for row in range(1, min(sheet.max_row, 100) + 1)]
        width = min(max(max((len(v) for v in values), default=8) + 2, 10), 48)
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, (datetime,)):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif hasattr(cell.value, "year") and hasattr(cell.value, "month") and hasattr(cell.value, "day"):
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(cell.value, int):
                cell.number_format = "0"


def _write_sheet(workbook: Workbook, name: str, headers: list[str], rows: list[tuple]) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"
            if name in {"超期与日期问题", "文号与机构问题", "外链问题", "元数据问题"}:
                cell.fill = ALERT_FILL
    _format_sheet(sheet)


def export_job(db: Database, job_id: int) -> Path:
    ensure_directories()
    with db.connect() as conn:
        job = conn.execute("SELECT * FROM scan_jobs WHERE id=?", (job_id,)).fetchone()
        if not job:
            raise ValueError("扫描任务不存在")
        findings = conn.execute(
            """SELECT f.*,d.district,d.source_id,d.source_site,d.title,d.url,d.issuing_agency,d.page_document_number,
            d.published_date,d.authored_date FROM rule_findings f JOIN policy_documents d ON d.id=f.document_id
            WHERE f.job_id=? ORDER BY d.district,d.id,f.id""", (job_id,)
        ).fetchall()
        links = conn.execute(
            """SELECT l.*,d.district,d.source_site,d.title,d.url AS document_url FROM link_checks l
            JOIN policy_documents d ON d.id=l.document_id WHERE l.job_id=? AND l.result!='ok'
            ORDER BY d.district,d.id,l.id""", (job_id,)
        ).fetchall()
        documents = conn.execute(
            """SELECT d.*,COALESCE(j.action,'processed') action,COALESCE(j.reason,'') reason,
            COALESCE(j.page_number,0) page_number,COALESCE(j.item_index,0) item_index,COUNT(f.id) finding_count
            ,COALESCE((SELECT raw_json FROM document_snapshots s WHERE s.document_id=d.id ORDER BY s.id DESC LIMIT 1),'{}') raw_json
            FROM policy_documents d LEFT JOIN scan_job_documents j ON d.id=j.document_id AND j.job_id=?
            LEFT JOIN rule_findings f ON f.document_id=d.id AND f.job_id=?
            WHERE j.id IS NOT NULL OR d.last_seen_job_id=?
            GROUP BY d.id,j.action,j.reason,j.page_number,j.item_index
            ORDER BY d.district,j.page_number,j.item_index""", (job_id, job_id, job_id)
        ).fetchall()
        item_results = conn.execute(
            """SELECT s.*,COALESCE(s.document_id,s.reused_document_id) AS effective_document_id
            FROM scan_item_results s WHERE s.job_id=?
            ORDER BY s.source_label,s.page_number,s.item_index""",
            (job_id,),
        ).fetchall()
        events = conn.execute(
            "SELECT * FROM job_events WHERE job_id=? ORDER BY id", (job_id,)
        ).fetchall()
        action_rows = conn.execute(
            """SELECT district,action,COUNT(*) count FROM scan_job_documents
            WHERE job_id=? GROUP BY district,action""", (job_id,)
        ).fetchall()

    workbook = Workbook()
    workbook.remove(workbook.active)
    findings_by_document: dict[int, list] = {}
    for finding in findings:
        findings_by_document.setdefault(int(finding["document_id"]), []).append(finding)
    categories: dict[tuple[str, str], int] = {}
    mapped_findings: set[int] = set()
    seen_source_findings: set[tuple[str, str, int]] = set()
    for item_result in item_results:
        document_id = item_result["effective_document_id"]
        if not document_id:
            continue
        for finding in findings_by_document.get(int(document_id), []):
            identity = (item_result["source_label"], item_result["url"], int(finding["id"]))
            if identity in seen_source_findings:
                continue
            seen_source_findings.add(identity)
            mapped_findings.add(int(finding["id"]))
            key = (item_result["source_label"], finding["category"])
            categories[key] = categories.get(key, 0) + 1
    for finding in findings:
        if int(finding["id"]) in mapped_findings:
            continue
        source = finding["source_site"] or finding["district"]
        key = (source, finding["category"])
        categories[key] = categories.get(key, 0) + 1
    coverage_label = "完整" if job["coverage_status"] == "complete" else "未完成"
    remaining = max(int(job["estimated_total"]) - int(job["examined_count"]), 0)
    selected_values = json.loads(job["districts_json"])
    sources = []
    for value in selected_values:
        try:
            label = resolve_target(value).label
        except ValueError:
            label = value
        if label not in sources:
            sources.append(label)
    totals_by_district = json.loads(job["total_by_district_json"] or "{}")
    examined_by_district = json.loads(job["examined_by_district_json"] or "{}")
    actions_by_district: dict[str, dict[str, int]] = {}
    for row in action_rows:
        actions_by_district.setdefault(row["district"], {})[row["action"]] = int(row["count"])
    if len(sources) == 1:
        source = sources[0]
        totals_by_district.setdefault(source, int(job["estimated_total"]))
        examined_by_district.setdefault(source, int(job["examined_count"]))
    summary_rows = []
    for source in sources:
        source_categories = [
            (category, count) for (name, category), count in sorted(categories.items()) if name == source
        ]
        if not source_categories:
            source_categories = [("无已发现问题", 0)]
        source_total = int(totals_by_district.get(source, 0))
        source_examined = int(examined_by_district.get(source, 0))
        source_actions = actions_by_district.get(source, {})
        source_processed = int(source_actions.get("processed", 0))
        source_skipped = int(source_actions.get("skipped", 0))
        source_remaining = max(source_total - source_examined, 0)
        for category, count in source_categories:
            summary_rows.append((
                source, category, count, job["status"], coverage_label, source_examined,
                source_processed, source_skipped, source_total, source_remaining,
                job["pause_reason"] or job["last_error"],
            ))
    _write_sheet(
        workbook, "问题汇总",
        ["扫描来源", "问题类型", "数量", "扫描状态", "扫描完整度", "已覆盖", "详情检查", "复用/跳过",
         "预计总数", "预计剩余", "暂停/未完成原因"], summary_rows,
    )

    date_rows = [(f["source_site"] or "市级平台", f["district"], f["source_id"], f["title"], _excel_date(f["authored_date"]),
                  _excel_date(f["published_date"]), _workday_count(f["detail"]),
                  "日期倒置" if f["rule_code"] == "DATE-002" else "超期", f["page_document_number"],
                  "页面", f["rule_code"], f["detail"], f["review_status"], f["url"])
                 for f in findings if f["category"] == "日期问题"]
    _write_sheet(workbook, "超期与日期问题", ["栏目名称", "区县", "文件ID", "标题", "成文日期", "发布日期", "工作日差", "时效状态", "文号", "文号来源", "规则", "问题详情", "复核状态", "链接"], date_rows)

    doc_rows = []
    for finding in findings:
        if finding["category"] not in {"文号问题", "机构问题"}:
            continue
        number_type, number_year, number_index, combined = _document_number_parts(finding["page_document_number"])
        doc_rows.append((finding["source_site"] or "市级平台", finding["district"], finding["source_id"], finding["title"],
                         _excel_date(finding["authored_date"]), number_type, number_year, number_index, combined,
                         "页面", finding["issuing_agency"], finding["body_value"], finding["rule_code"],
                         finding["detail"], finding["evidence"], finding["review_status"], finding["url"]))
    _write_sheet(workbook, "文号与机构问题", ["栏目名称", "区县", "文件ID", "标题", "成文日期", "文号类型", "文号年份", "文号编号", "组合文号", "文号来源", "发文机构", "正文值", "问题类型", "问题详情", "证据", "复核状态", "链接"], doc_rows)

    link_rows = [(r["source_site"] or "市级平台", r["district"], r["title"], r["link_kind"], r["original_url"], r["final_url"],
                  r["status_code"], r["result"], r["error_type"], r["redirect_chain_json"], r["page_title"],
                  r["checked_at"], r["document_url"]) for r in links]
    _write_sheet(workbook, "外链问题", ["来源", "区县", "文件标题", "关联类型", "原始URL", "最终URL", "状态码", "检查结果", "错误类型", "重定向链", "页面标题", "检查时间", "文件链接"], link_rows)

    status_names = {
        "checked_complete": "表头完整，已检查",
        "checked_incomplete": "表头不完整，已记录问题",
        "no_header_pass": "无表头 PASS",
        "reused_baseline_detail": "基线复用（有表头）",
        "reused_baseline_no_header": "基线复用（无表头 PASS）",
        "reused_current_detail": "当期 URL 复用（有表头）",
        "reused_current_no_header": "当期 URL 复用（无表头 PASS）",
        "exception": "详情异常，待复测",
    }
    all_rows = []
    metadata_by_url: dict[str, tuple[int, tuple]] = {}
    for item_result in item_results:
        document_id = item_result["effective_document_id"]
        doc_findings = findings_by_document.get(int(document_id), []) if document_id else []
        issue_codes = "、".join(finding["rule_code"] for finding in doc_findings)
        issue_details = "；".join(finding["detail"] for finding in doc_findings)
        try:
            missing_fields = json.loads(item_result["missing_fields_json"] or "[]")
        except json.JSONDecodeError:
            missing_fields = ["缺失字段记录格式异常"]
        authored = _excel_date(item_result["authored_date"])
        published = _excel_date(item_result["published_date"])
        workdays, time_status = _timeliness(authored, published)
        number_type, number_year, number_index, combined = _document_number_parts(item_result["page_document_number"])
        header_state = "无表头" if not item_result["header_detected"] else "表头不完整" if missing_fields else "表头完整"
        all_rows.append((
            item_result["source_label"], item_result["channel_id"], item_result["page_number"],
            item_result["item_index"] + 1, _excel_date(item_result["listed_date"]), item_result["title"],
            header_state, item_result["source_id"], item_result["topic_category"], item_result["disclosure_attribute"],
            authored, item_result["page_document_number"], published, item_result["issuing_agency"],
            "、".join(missing_fields), workdays, time_status, number_type, number_year, number_index, combined,
            issue_codes, issue_details, len(doc_findings), status_names.get(item_result["detail_status"], item_result["detail_status"]),
            item_result["baseline_job_id"] or "", item_result["reason"], item_result["url"],
        ))
        if missing_fields:
            metadata_row = (
                item_result["source_label"], item_result["channel_id"], item_result["title"], item_result["source_id"],
                "、".join(missing_fields), item_result["detail_status"], item_result["reason"], item_result["url"],
            )
            priority = 2 if item_result["detail_status"] == "checked_incomplete" else 1
            current = metadata_by_url.get(item_result["url"])
            if current is None or priority > current[0]:
                metadata_by_url[item_result["url"]] = (priority, metadata_row)
    metadata_rows = [row for _priority, row in metadata_by_url.values()]
    _write_sheet(
        workbook, "全量明细",
        ["来源 Sheet", "栏目 ID", "列表页码", "列表序号", "列表发布日期", "标题", "表头状态", "索引号", "主题分类",
         "公开属性", "成文日期", "发文字号", "发布日期", "公开主体", "缺失/无效字段", "工作日差", "时效状态",
         "文号类型", "文号年份", "文号编号", "组合文号", "问题类型", "问题详情", "问题数", "处理结果",
         "基线任务 ID", "处理说明", "链接"],
        all_rows,
    )
    _write_sheet(
        workbook, "元数据问题",
        ["来源 Sheet", "栏目 ID", "标题", "索引号", "缺失/无效字段", "处理结果", "处理说明", "链接"],
        metadata_rows,
    )

    run_rows = [(job["id"], "任务汇总", job["mode"], job["status"], coverage_label, job["completion_kind"],
                 job["created_at"], job["started_at"], job["finished_at"], job["examined_count"],
                 job["processed_count"], job["skipped_count"], job["estimated_total"], remaining,
                 job["access_count"], job["retry_count"], job["rest_count"], job["resumed_count"],
                 job["current_district"], job["current_page"], job["pause_reason"], job["last_error"],
                 job["safety_json"], "", "")]
    run_rows.extend((job["id"], event["event_type"], job["mode"], job["status"], coverage_label,
                     job["completion_kind"], event["created_at"], "", "", "", "", "", "", "", "", "", "", "",
                     "", "", event["message"], "", event["details_json"], event["url"], event["id"]) for event in events)
    _write_sheet(workbook, "扫描运行记录", ["任务ID", "记录类型", "模式", "状态", "扫描完整度", "结束类型", "记录时间", "开始时间", "结束时间", "已覆盖", "详情检查", "复用/跳过", "预计总数", "预计剩余", "访问次数", "重试次数", "强休息次数", "恢复次数", "当前区县", "当前页", "说明", "最后错误", "参数/详情", "URL", "事件ID"], run_rows)

    output = EXPORT_DIR / f"政策巡检结果_任务{job_id}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    workbook.save(output)
    return output
