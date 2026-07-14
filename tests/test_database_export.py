import sqlite3
from datetime import date, datetime

from openpyxl import load_workbook

from app.db import Database
from app.domain import Finding, PolicyListItem, PolicyRecord
from app.exporter import export_job
from app.repository import Repository


def test_database_and_excel_export(tmp_path, monkeypatch):
    db = Database(tmp_path / "test.db")
    db.initialize()
    job_id = db.create_job(["putuo_government"], "full", {"min_delay_seconds": 5})
    db.update_job(job_id, status="completed", coverage_status="complete", completion_kind="full",
                  processed_count=1, examined_count=1, estimated_total=1, finding_count=1)
    record = PolicyRecord(
        district="普陀区", source_id="10001", title="测试文件", url="https://example.test/10001.html",
        issuing_agency="上海市普陀区人民政府", page_document_number="普府〔2026〕1号",
        authored_date=date(2026, 1, 1), published_date=date(2026, 2, 1), body_text="测试正文",
    )
    repository = Repository(db)
    document_id = repository.save_record(
        job_id, record, [Finding("DATE-001", "日期问题", "high", "confirmed", "相隔 22 个工作日，超过20个工作日")]
    )
    repository.record_job_document(
        job_id, document_id, "区级网站·普陀区·区政府文件", 1, 0, "processed"
    )
    repository.record_scan_item(
        job_id,
        PolicyListItem(
            district="普陀区", page_number=1, item_index=0, title="测试文件",
            url="https://example.test/10001.html", published_date=date(2026, 2, 1),
            source_site="区级网站·普陀区·区政府文件", source_key="putuo_government", source_channel_id="3",
        ),
        detail_status="checked_complete", header_detected=True, source_id="10001",
        authored_date=date(2026, 1, 1), page_document_number="普府〔2026〕1号",
        published_date=date(2026, 2, 1), issuing_agency="上海市普陀区人民政府", document_id=document_id,
    )
    repository.save_link_check(
        job_id,
        document_id,
        {
            "kind": "政策解读",
            "url": "https://example.test/read-missing",
            "final_url": "https://example.test/read-missing",
            "status_code": 404,
            "result": "broken",
            "error_type": "http_error",
            "source_area": "详情页侧栏",
            "link_text": "政策解读",
            "source_page_url": record.url,
            "evidence": "HTTP 404",
        },
    )
    repository.save_link_check(
        job_id,
        document_id,
        {
            "kind": "阅办联动",
            "url": "https://example.test/restricted",
            "final_url": "https://example.test/restricted",
            "status_code": 403,
            "result": "review_required",
            "error_type": "access_restricted",
            "source_area": "列表页",
            "link_text": "阅办联动",
            "source_page_url": "https://example.test/list",
            "review_status": "manual_review",
            "evidence": "访问限制，待人工复核",
        },
    )
    repository.save_link_check(
        job_id,
        document_id,
        {
            "kind": "阅办联动",
            "url": "https://api.example.test/legacy-hidden",
            "final_url": "https://api.example.test/legacy-hidden",
            "status_code": 404,
            "result": "broken",
            "error_type": "http_error",
        },
    )
    monkeypatch.setattr("app.exporter.EXPORT_DIR", tmp_path)
    output = export_job(db, job_id)
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["问题汇总", "超期与日期问题", "文号与机构问题", "外链问题", "全量明细", "元数据问题", "扫描运行记录"]
    assert workbook["全量明细"]["F2"].value == "测试文件"
    assert workbook["全量明细"]["AB2"].hyperlink.target == "https://example.test/10001.html"
    assert workbook["超期与日期问题"]["E2"].value == datetime(2026, 1, 1)
    assert workbook["超期与日期问题"]["G2"].value == 22
    assert workbook["问题汇总"]["E2"].value == "完整"
    assert workbook["问题汇总"]["F2"].value == 1
    assert workbook["问题汇总"]["G2"].value == 1
    assert workbook["问题汇总"]["A2"].value == "区级网站·普陀区·区政府文件"
    assert workbook["问题汇总"]["B2"].value == "日期问题"
    assert workbook["问题汇总"]["C2"].value == 1
    assert workbook["外链问题"].max_row == 2
    assert workbook["外链问题"]["E2"].value == "详情页侧栏"
    assert workbook["外链问题"]["F2"].value == "政策解读"
    assert workbook["外链问题"]["H2"].hyperlink.target == "https://example.test/read-missing"


def test_old_link_checks_schema_migrates_before_occurrence_index(tmp_path):
    path = tmp_path / "legacy.db"
    with sqlite3.connect(path) as conn:
        conn.executescript(
            """
            CREATE TABLE link_checks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id INTEGER NOT NULL,
                document_id INTEGER NOT NULL,
                link_kind TEXT NOT NULL,
                original_url TEXT NOT NULL,
                final_url TEXT NOT NULL DEFAULT '',
                status_code INTEGER,
                result TEXT NOT NULL,
                error_type TEXT NOT NULL DEFAULT '',
                page_title TEXT NOT NULL DEFAULT '',
                checked_at TEXT NOT NULL
            );
            CREATE UNIQUE INDEX uq_link_check_per_job_url
                ON link_checks(job_id, document_id, original_url);
            INSERT INTO link_checks(
                job_id,document_id,link_kind,original_url,result,checked_at
            ) VALUES(1,1,'阅办联动','https://api.example.test/hidden','broken','2026-07-14');
            CREATE TABLE scan_item_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id INTEGER NOT NULL,
                target_key TEXT NOT NULL,
                source_label TEXT NOT NULL,
                channel_id TEXT NOT NULL DEFAULT '',
                page_number INTEGER NOT NULL,
                item_index INTEGER NOT NULL,
                title TEXT NOT NULL,
                url TEXT NOT NULL,
                listed_date TEXT,
                detail_status TEXT NOT NULL,
                header_detected INTEGER NOT NULL DEFAULT 0,
                source_id TEXT NOT NULL DEFAULT '',
                topic_category TEXT NOT NULL DEFAULT '',
                disclosure_attribute TEXT NOT NULL DEFAULT '',
                authored_date TEXT,
                page_document_number TEXT NOT NULL DEFAULT '',
                published_date TEXT,
                issuing_agency TEXT NOT NULL DEFAULT '',
                missing_fields_json TEXT NOT NULL DEFAULT '[]',
                document_id INTEGER,
                reused_document_id INTEGER,
                baseline_job_id INTEGER,
                reason TEXT NOT NULL DEFAULT '',
                checked_at TEXT NOT NULL,
                UNIQUE(job_id,target_key,page_number,item_index)
            );
            INSERT INTO scan_item_results(
                job_id,target_key,source_label,page_number,item_index,title,url,detail_status,checked_at
            ) VALUES(1,'municipal_putuo','legacy',1,0,'legacy','https://example.test/legacy',
                     'checked_complete','2026-07-14');
            """
        )

    db = Database(path)
    db.initialize()
    db.initialize()

    with db.connect() as conn:
        row = conn.execute(
            "SELECT source_area,source_page_url,visible,review_status FROM link_checks"
        ).fetchone()
        indexes = {item["name"] for item in conn.execute("PRAGMA index_list(link_checks)")}
        old_item_version = conn.execute(
            "SELECT link_check_version FROM scan_item_results WHERE title='legacy'"
        ).fetchone()[0]
    assert row["source_area"] == row["source_page_url"] == ""
    assert row["visible"] == 0
    assert row["review_status"] == "legacy_hidden"
    assert "uq_link_check_per_job_url" not in indexes
    assert "uq_link_check_per_occurrence" in indexes
    assert old_item_version == 0


def test_copy_baseline_links_keeps_page_evidence_but_skips_legacy_hidden(tmp_path):
    db = Database(tmp_path / "copy-links.db")
    db.initialize()
    baseline_job = db.create_job(["普陀区"], "full", {})
    current_job = db.create_job(["普陀区"], "incremental", {}, baseline_job_id=baseline_job)
    repository = Repository(db)
    document_id = repository.save_record(
        baseline_job,
        PolicyRecord("普陀区", "测试", "https://example.test/policy"),
        [],
    )
    common = {
        "kind": "政策解读",
        "final_url": "https://example.test/link",
        "status_code": 200,
        "result": "ok",
        "source_area": "详情页侧栏",
        "source_page_url": "https://example.test/policy",
    }
    repository.save_link_check(
        baseline_job, document_id,
        {**common, "url": "https://example.test/confirmed", "link_text": "图文解读"},
    )
    repository.save_link_check(
        baseline_job, document_id,
        {
            **common,
            "url": "https://example.test/manual",
            "link_text": "视频解读",
            "result": "review_required",
            "review_status": "manual_review",
        },
    )
    repository.save_link_check(
        baseline_job, document_id,
        {
            **common,
            "url": "https://api.example.test/legacy",
            "link_text": "接口关系",
            "visible": False,
            "review_status": "legacy_hidden",
        },
    )

    repository.copy_baseline_link_checks(baseline_job, current_job, document_id)

    with db.connect() as conn:
        rows = conn.execute(
            "SELECT original_url,review_status FROM link_checks WHERE job_id=? ORDER BY original_url",
            (current_job,),
        ).fetchall()
    assert [(row["original_url"], row["review_status"]) for row in rows] == [
        ("https://example.test/confirmed", "confirmed"),
        ("https://example.test/manual", "manual_review"),
    ]


def test_export_deduplicates_metadata_problem_sheet_by_url(tmp_path, monkeypatch):
    db = Database(tmp_path / "metadata.db")
    db.initialize()
    job_id = db.create_job(["putuo_government"], "full", {})
    db.update_job(
        job_id, status="completed", coverage_status="complete", completion_kind="full",
        processed_count=1, skipped_count=1, examined_count=2, estimated_total=2, finding_count=1,
        total_by_district_json='{"区级网站·普陀区·区政府文件": 2}',
        examined_by_district_json='{"区级网站·普陀区·区政府文件": 2}',
    )
    url = "https://www.shpt.gov.cn/zhengwu/test/1.html"
    source_label = "区级网站·普陀区·区政府文件"
    record = PolicyRecord(
        district="普陀区", title="元数据测试", url=url, source_site=source_label,
        source_id="SY1", published_date=date(2026, 7, 1), authored_date=date(2026, 7, 1),
        header_detected=True, missing_metadata_fields=["主题分类"],
    )
    repository = Repository(db)
    document_id = repository.save_record(
        job_id, record,
        [Finding("META-001", "元数据问题", "high", "confirmed", "详情页表头字段缺失：主题分类")],
    )
    repository.record_job_document(job_id, document_id, source_label, 1, 0, "processed")
    first = PolicyListItem(
        "普陀区", 1, 0, "元数据测试", url, date(2026, 7, 1),
        source_site=source_label, source_key="putuo_government", source_channel_id="3",
    )
    duplicate = PolicyListItem(
        "普陀区", 2, 0, "元数据测试", url, date(2026, 7, 1),
        source_site=source_label, source_key="putuo_government", source_channel_id="3",
    )
    repository.record_scan_item(
        job_id, first, detail_status="checked_incomplete", header_detected=True,
        source_id="SY1", authored_date=date(2026, 7, 1), published_date=date(2026, 7, 1),
        missing_fields=["主题分类"], document_id=document_id, reason="实际详情检查",
    )
    repository.record_scan_item(
        job_id, duplicate, detail_status="reused_current_detail", header_detected=True,
        source_id="SY1", authored_date=date(2026, 7, 1), published_date=date(2026, 7, 1),
        missing_fields=["主题分类"], document_id=document_id, reused_document_id=document_id,
        reason="同一 URL 复用",
    )

    monkeypatch.setattr("app.exporter.EXPORT_DIR", tmp_path)
    workbook = load_workbook(export_job(db, job_id))

    assert workbook["全量明细"].max_row == 3
    assert workbook["元数据问题"].max_row == 2
    assert workbook["元数据问题"]["F2"].value == "checked_incomplete"
    assert workbook["问题汇总"]["A2"].value == source_label
    assert workbook["问题汇总"]["B2"].value == "元数据问题"
    assert workbook["问题汇总"]["C2"].value == 1
